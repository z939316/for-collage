import urllib.request as req
import requests
from openpyxl import workbook
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import openpyxl
from openpyxl.utils import get_column_letter
import time
import datetime
import re
#微星VENTUS=[]
#微星GAMING=[]
#技嘉D6D6=[]
#技嘉OC 1725OC=[]
#技嘉20.5cmWINDFORCE=[]
#EVGASC=[]
#ZOTACTwin=[]
p=[0,0,0,0,0,0,0]
url="https://www.coolpc.com.tw/evaluate.php"
request=req.Request(url,headers={
    "User-Agent":"Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/88.0.4324.150 Safari/537.36"
})
data=requests.get(url)
file=open('C:/Users/USER/Desktop/boot/GPU.txt',mode="w",encoding="utf-8")
file.write(data.text)

def add(n):
    n=int(n)
    if "VENTUS" in line:
        p[0]=n
    elif "GAMING" and "微星" in line:
        p[1]=n
    elif "D6" in line:
        p[2]=n
    elif "OC" and "4G(1725MHz/17.2cm/單風扇)" in line:
        p[3]=n
    elif "WINDFORCE" in line:
        p[4]=n
    elif "SC" in line:
        p[5]=n
    elif "Twin" in line:
        p[6]=n

with open("C:/Users/USER/Desktop/boot/GPU.txt",mode="r",encoding="utf-8") as file:
    for line in file:
        if "GTX1650 SUPER" in line:
            #print(line)
            line=list(map(str,line.split()))
            
            for i in range(len(line)):
                if "$" in line[i]:
                    if line[i+1].isdigit():
                        add(line[i+1])
                        #print(line)
                    else:
                        line[i]=line[i].replace("$",'')
                        line[i]=line[i].replace("4G(1755MHz/20.2cm)參考價","")
                        add(line[i])
                        #print(line[i])
wb = load_workbook('C:/Users/USER/Desktop/boot/GPU.xlsx')
sheet = wb['工作表1']
a=re.split('-| ',str(datetime.datetime.now()))
month=int(a[1])
day=int(a[2])
count=month*30+day-254
sheet.cell(row=1, column=count, value=a[1]+'/'+a[2])
f=0
for i in range(7):
    sheet.cell(row=i+2, column=count, value=p[i])
    if p[i] < sheet.cell(i+2, count-1).value:
        print(sheet.cell(i+2, 1).value+"降價 "+str(sheet.cell(i+2, count-1).value-p[i])+'元')
        f=1
    if p[i] > sheet.cell(i+2, count-1).value:
        print(sheet.cell(i+2, 1).value+"漲價 "+str(p[i]-sheet.cell(i+2, count-1).value)+'元')
        f=1
    #sheet.cell(i+2,count)=p[i]
wb.save('GPU.xlsx')
if f==0:
    print("今日顯卡價格無異動")







